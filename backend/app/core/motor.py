"""
FinAnalytics — Motor de Análisis Financiero
============================================
Implementa la metodología real usada en estudios de sector para Colombia
(Decreto 1082/2015, Colombia Compra Eficiente), tal como se documenta
para análisis de sector bajo Decreto 1082/2015 (Colombia Compra Eficiente).

Metodología objetiva (Modo A):
  1. Filtrar empresas del sector por CIIU
  2. Seleccionar muestra estratificada proporcional por departamento (~3%)
  3. Calcular IL, IE, RCI, RP, RA sobre cada empresa de la muestra
  4. Construir tabla de frecuencias por rangos para cada índice
  5. Calcular Hi acumulada descendente (IL,RCI,RP,RA) o ascendente (IE)
  6. Identificar el rango donde Hi >= umbral (~59-60%)
  7. El índice recomendado = límite superior de ese rango
  8. Generar narrativa automática ("el X% de empresas tiene índice >= Y...")

Metodología por objetivo de empresa (Modo B):
  Misma distribución, pero ubica el índice objetivo del usuario dentro
  de ella. Calcula cuánto % del sector cumple ese objetivo y ofrece
  una opción inclusiva (rango concentrado) y una exigente (objetivo usuario).
"""

from __future__ import annotations
import io, logging
from dataclasses import dataclass, field
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
#  COLUMNAS SIIS
# ─────────────────────────────────────────────────────────────────────────────
COL = {
    "nit":        "NIT",
    "nombre":     "Razón social de la sociedad",
    "ciiu":       "Clasificación Industrial Internacional Uniforme Versión 4 A.C (CIIU)",
    "dpto":       "Departamento de la dirección del domicilio",
    "ciudad":     "Ciudad de la dirección del domicilio",
    "act_corr":   "Activos corrientes totales (CurrentAssets)",
    "act_total":  "Total de activos (Assets)",
    "pas_corr":   "Pasivos corrientes totales (CurrentLiabilities)",
    "pas_total":  "Total pasivos (Liabilities)",
    "patrimonio": "Patrimonio total (Equity)",
    "util_op":    "Ganancia (pérdida) por actividades de operación (GananciaPerdidaPorActividadesDeOperacion)",
    "gastos_int": "Costos financieros (FinanceCosts)",
}

# ─────────────────────────────────────────────────────────────────────────────
#  RANGOS FIJOS POR ÍNDICE — metodología de análisis sectorial colombiano
#  Tupla: (etiqueta, lim_inferior, lim_superior)   None = abierto
# ─────────────────────────────────────────────────────────────────────────────
RANGOS = {
    "IL": [
        ("0,00 - 0,99",   0.00,  1.00),
        ("1,00 - 1,49",   1.00,  1.50),
        ("1,50 - 1,99",   1.50,  2.00),
        ("2,00 - 2,49",   2.00,  2.50),
        ("2,50 - 2,99",   2.50,  3.00),
        ("3,00 - 3,49",   3.00,  3.50),
        ("3,50 - 3,99",   3.50,  4.00),
        ("4,00 - 4,49",   4.00,  4.50),
        ("4,50 y mayor",  4.50,  None),
    ],
    "IE": [
        ("Menor a 10%",   None,  0.10),
        ("10% - 20%",     0.10,  0.20),
        ("20% - 30%",     0.20,  0.30),
        ("30% - 40%",     0.30,  0.40),
        ("40% - 50%",     0.40,  0.50),
        ("50% - 60%",     0.50,  0.60),
        ("60% - 70%",     0.60,  0.70),
        ("70% - 80%",     0.70,  0.80),
        ("Mayor a 80%",   0.80,  None),
    ],
    "RCI": [
        ("Menor a 0",     None,  0.00),
        ("0 - 1",         0.00,  1.00),
        ("1 - 2",         1.00,  2.00),
        ("2 - 3",         2.00,  3.00),
        ("3 - 4",         3.00,  4.00),
        ("4 - 5",         4.00,  5.00),
        ("Mayor a 5",     5.00,  None),
    ],
    "RP": [
        ("Menor a 0%",    None,  0.00),
        ("0% - 1%",       0.00,  0.01),
        ("1% - 2%",       0.01,  0.02),
        ("2% - 3%",       0.02,  0.03),
        ("3% - 4%",       0.03,  0.04),
        ("4% - 5%",       0.04,  0.05),
        ("Mayor a 5%",    0.05,  None),
    ],
    "RA": [
        ("Menor a 0%",    None,  0.00),
        ("0% - 1%",       0.00,  0.01),
        ("1% - 2%",       0.01,  0.02),
        ("2% - 3%",       0.02,  0.03),
        ("3% - 4%",       0.03,  0.04),
        ("4% - 5%",       0.04,  0.05),
        ("Mayor a 5%",    0.05,  None),
    ],
}

# "mayor" → Hi descendente (mayor valor = mejor): IL, RCI, RP, RA
# "menor" → Hi ascendente (menor valor = mejor): IE
DIRECCION = {"IL": "mayor", "IE": "menor", "RCI": "mayor", "RP": "mayor", "RA": "mayor"}

# Umbral de concentración Hi ≥ 59% (como usa el documento)
HI_UMBRAL = 0.59

NARRATIVA = {
    "IL":  {"titulo": "ÍNDICE DE LIQUIDEZ (RAZÓN CORRIENTE)",
            "intro": "Los indicadores de liquidez miden la capacidad de la empresa para afrontar sus obligaciones de corto plazo. La razón corriente verifica si los activos corrientes cubren los pasivos corrientes.",
            "formula": "IL = Activo Corriente / Pasivo Corriente"},
    "IE":  {"titulo": "ÍNDICE DE ENDEUDAMIENTO (RAZÓN DE ENDEUDAMIENTO)",
            "intro": "Mide la proporción en que el proponente está financiado por terceros. A mayor índice, mayor dependencia de recursos externos y mayor probabilidad de incumplimiento.",
            "formula": "IE = Total Pasivos / Total Activos"},
    "RCI": {"titulo": "RAZÓN DE COBERTURA DE INTERESES",
            "intro": "Refleja la capacidad de la empresa para cumplir obligaciones financieras. A mayor cobertura, menor probabilidad de incumplimiento de obligaciones financieras.",
            "formula": "RCI = Utilidad Operacional / Costos Financieros"},
    "RP":  {"titulo": "RENTABILIDAD DEL PATRIMONIO",
            "intro": "Mide la capacidad de generación de utilidad por cada peso invertido en el patrimonio. A mayor rentabilidad, mejor capacidad organizacional de la empresa.",
            "formula": "RP = Utilidad Operacional / Patrimonio"},
    "RA":  {"titulo": "RENTABILIDAD DEL ACTIVO",
            "intro": "Mide la eficiencia en la generación de utilidad por cada peso invertido en activos. A mayor rentabilidad del activo, mayor eficiencia operacional.",
            "formula": "RA = Utilidad Operacional / Total Activos"},
}

# ─────────────────────────────────────────────────────────────────────────────
#  SCHEMAS
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ConfigAnalisis:
    ciius: list[str]
    modo: str = "A"                            # "A" objetivo | "B" empresa
    indices_empresa: dict[str, float] | None = None
    porcentaje_muestra: float = 0.03
    hi_umbral: float = HI_UMBRAL
    seed: int = 42
    año_datos: int | None = None
    # Modo B: número fijo de empresas a seleccionar (si None usa porcentaje_muestra)
    n_empresas_b: int | None = None

    def validar(self):
        if not self.ciius:
            raise ValueError("Seleccione al menos un código CIIU.")
        if self.modo not in ("A", "B"):
            raise ValueError("El modo debe ser 'A' o 'B'.")
        if self.modo == "B":
            if not self.indices_empresa:
                raise ValueError("Modo B requiere indices_empresa.")
            faltantes = [k for k in RANGOS if k not in self.indices_empresa]
            if faltantes:
                raise ValueError(f"Faltan índices en indices_empresa: {faltantes}")


@dataclass
class TablaFrecuencia:
    indice: str
    filas: list[dict]           # [{Rango, Frecuencia, %, Hi, Empresas acumuladas, _meta}]
    n_total: int
    rango_concentrado: str
    indice_recomendado: float
    hi_concentrado: float
    pct_cumplen: float
    narrativa: str
    # Umbral sugerido por segunda derivada (punto de inflexión real de la curva Hi)
    umbral_sugerido: float | None = None       # ej: 0.59
    indice_recomendado_sugerido: float | None = None  # índice con umbral sugerido
    rango_concentrado_sugerido: str | None = None


@dataclass
class ResultadoIndice:
    tabla: TablaFrecuencia
    objetivo_empresa: float | None = None
    rango_objetivo: str | None = None
    pct_sector_cumple: float | None = None
    indice_inclusivo: float | None = None
    narrativa_b: str | None = None
    # Modo B nuevo: métricas de cercanía al objetivo
    promedio_grupo: float | None = None        # promedio del índice en el grupo seleccionado
    diferencia_abs: float | None = None        # |promedio - objetivo|
    diferencia_pct: float | None = None        # diferencia relativa al objetivo
    cumple_objetivo: bool | None = None        # promedio <= objetivo (IE) o >= objetivo (resto)


@dataclass
class ResultadoAnalisis:
    config: ConfigAnalisis
    tabla_poblacion: pd.DataFrame
    tabla_muestra: pd.DataFrame
    n_poblacion: int
    n_muestra: int
    resultados: dict[str, ResultadoIndice]
    df_muestra: pd.DataFrame
    advertencias: list[str] = field(default_factory=list)
    # Modo B: resumen global de cercanía al objetivo
    resumen_modo_b: dict | None = None

    def a_dict(self) -> dict[str, Any]:
        return {
            "config": {
                "ciius": self.config.ciius,
                "modo": self.config.modo,
                "indices_empresa": self.config.indices_empresa,
                "porcentaje_muestra": self.config.porcentaje_muestra,
                "n_empresas_b": self.config.n_empresas_b,
                "hi_umbral": self.config.hi_umbral,
                "año_datos": self.config.año_datos,
            },
            "resumen": {
                "n_poblacion": self.n_poblacion,
                "n_muestra": self.n_muestra,
                "advertencias": self.advertencias,
                "resumen_modo_b": self.resumen_modo_b,
            },
            "geografico": {
                "poblacion": self.tabla_poblacion.to_dict("records"),
                "muestra": self.tabla_muestra.to_dict("records"),
            },
            "indices": {
                idx: {
                    "tabla_frecuencias": [
                        {k: v for k, v in f.items() if not k.startswith("_")}
                        for f in res.tabla.filas
                    ],
                    "rango_concentrado": res.tabla.rango_concentrado,
                    "indice_recomendado": res.tabla.indice_recomendado,
                    "hi_concentrado": res.tabla.hi_concentrado,
                    "umbral_sugerido": res.tabla.umbral_sugerido,
                    "indice_recomendado_sugerido": res.tabla.indice_recomendado_sugerido,
                    "rango_concentrado_sugerido": res.tabla.rango_concentrado_sugerido,
                    "pct_cumplen": res.tabla.pct_cumplen,
                    "narrativa": res.tabla.narrativa,
                    "objetivo_empresa": res.objetivo_empresa,
                    "rango_objetivo": res.rango_objetivo,
                    "pct_sector_cumple": res.pct_sector_cumple,
                    "indice_inclusivo": res.indice_inclusivo,
                    "narrativa_b": res.narrativa_b,
                    "promedio_grupo": res.promedio_grupo,
                    "diferencia_abs": res.diferencia_abs,
                    "diferencia_pct": res.diferencia_pct,
                    "cumple_objetivo": res.cumple_objetivo,
                }
                for idx, res in self.resultados.items()
            },
        }


# ─────────────────────────────────────────────────────────────────────────────
#  MOTOR PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

class MotorAnalisis:
    def __init__(self, df: pd.DataFrame, col: dict | None = None):
        self.df_raw = df.copy()
        self.col = col or COL

    def ejecutar(self, config: ConfigAnalisis) -> ResultadoAnalisis:
        config.validar()
        advertencias = []

        df_sector = self._filtrar_ciiu(self.df_raw, config.ciius)
        n_poblacion = len(df_sector)
        if n_poblacion == 0:
            raise ValueError(f"No hay empresas con los CIIUs {config.ciius}.")

        tabla_pob = self._tabla_geo(df_sector)
        df_muestra = self._seleccionar_muestra(df_sector, config.porcentaje_muestra, config.seed)
        n_muestra = len(df_muestra)
        tabla_mue = self._tabla_geo(df_muestra)

        if n_muestra < 30:
            if n_muestra < 10:
                advertencias.append(
                    f"⚠ Muestra muy pequeña ({n_muestra} empresas). "
                    "Los resultados pueden no ser representativos del sector. "
                    "Se recomienda ampliar los CIIUs seleccionados o aumentar el porcentaje de muestra."
                )
            else:
                advertencias.append(
                    f"⚠ Muestra pequeña ({n_muestra} empresas). "
                    "Para mayor precisión estadística se recomienda una muestra de al menos 30 empresas. "
                    "Considere ampliar los CIIUs seleccionados o el porcentaje de muestra."
                )

        df_muestra = self._calcular_indices(df_muestra)

        resultados = {}
        resumen_modo_b = None

        if config.modo == "B" and config.indices_empresa:
            # ── MODO B: selección por optimización ────────────────────────
            # Determinar N objetivo
            n_obj_b = config.n_empresas_b if config.n_empresas_b else max(10, round(n_poblacion * config.porcentaje_muestra))
            # Seleccionar el grupo óptimo cuyo promedio minimiza distancia a objetivos
            df_grupo, resumen_modo_b = self._seleccionar_grupo_modo_b(
                df_sector, config.indices_empresa, n_obj_b, config.seed
            )
            df_muestra = df_grupo  # el grupo seleccionado ES la muestra del Modo B
            n_muestra = len(df_muestra)

            for idx in RANGOS:
                serie = df_muestra[idx].dropna()
                tabla = self._tabla_frecuencias(idx, serie, config.hi_umbral)
                res = self._modo_b_metricas(tabla, config.indices_empresa[idx], idx, serie, resumen_modo_b)
                resultados[idx] = res
        else:
            # ── MODO A: distribución del sector ───────────────────────────
            for idx in RANGOS:
                serie = df_muestra[idx].dropna()
                tabla = self._tabla_frecuencias(idx, serie, config.hi_umbral)
                resultados[idx] = ResultadoIndice(tabla=tabla)

        return ResultadoAnalisis(
            config=config,
            tabla_poblacion=tabla_pob,
            tabla_muestra=tabla_mue,
            n_poblacion=n_poblacion,
            n_muestra=n_muestra,
            resultados=resultados,
            df_muestra=df_muestra,
            advertencias=advertencias,
            resumen_modo_b=resumen_modo_b,
        )

    # ── FILTRO CIIU ────────────────────────────────────────────────────────
    def _filtrar_ciiu(self, df, ciius):
        col = self.col["ciiu"]
        if col not in df.columns:
            return df.copy()
        mask = df[col].astype(str).str.strip().str[:5].isin(ciius)
        return df[mask].copy().reset_index(drop=True)

    # ── DISTRIBUCIÓN GEOGRÁFICA ────────────────────────────────────────────
    def _tabla_geo(self, df: pd.DataFrame) -> pd.DataFrame:
        col = self.col.get("dpto", "")
        if col not in df.columns:
            return pd.DataFrame(columns=["Departamento", "Número de Empresas"])
        conteo = df[col].value_counts().reset_index()
        conteo.columns = ["Departamento", "Número de Empresas"]
        total = pd.DataFrame([{"Departamento": "Total general", "Número de Empresas": len(df)}])
        return pd.concat([conteo, total], ignore_index=True)

    # ── MUESTRA ESTRATIFICADA POR DEPTO ────────────────────────────────────
    def _seleccionar_muestra(self, df, pct, seed):
        col_dpto = self.col.get("dpto", "")
        n_obj = max(1, round(len(df) * pct))
        if col_dpto not in df.columns:
            return df.sample(n=min(n_obj, len(df)), random_state=seed)
        grupos = df.groupby(col_dpto, group_keys=False)
        total = len(df)
        sel = []
        for _, g in grupos:
            n_g = max(1, round(n_obj * len(g) / total))
            n_g = min(n_g, len(g))
            sel.append(g.sample(n=n_g, random_state=seed))
        return pd.concat(sel).reset_index(drop=True)

    # ── CÁLCULO DE INDICADORES ─────────────────────────────────────────────
    def _calcular_indices(self, df):
        df = df.copy()
        eps = 1e-9

        def s(key):
            nombre = self.col.get(key, key)
            if nombre in df.columns:
                return pd.to_numeric(df[nombre], errors="coerce")
            return pd.Series(np.nan, index=df.index)

        ac = s("act_corr"); pc = s("pas_corr")
        at = s("act_total"); pt = s("pas_total")
        pa = s("patrimonio"); uo = s("util_op"); gi = s("gastos_int")

        df["CT"]  = ac - pc
        df["IL"]  = ac / pc.replace(0, eps)
        df["IE"]  = pt / at.replace(0, eps)

        # RCI: si costos financieros es 0 o casi 0, el índice no aplica (NaN)
        # No tiene sentido calcular cobertura de intereses si no hay intereses
        gi_safe = gi.copy()
        gi_safe[gi_safe.abs() < 1.0] = np.nan   # menos de $1.000 en costos fin. → NaN
        df["RCI"] = uo / gi_safe

        # RP y RA: si denominador es casi 0 o negativo extremo, NaN
        pa_safe = pa.copy()
        pa_safe[pa_safe.abs() < 1.0] = np.nan
        df["RP"]  = uo / pa_safe

        df["RA"]  = uo / at.replace(0, eps)

        for idx in ["IL", "IE", "RCI", "RP", "RA"]:
            df[idx] = df[idx].replace([np.inf, -np.inf], np.nan)
        return df

    # ── TABLA DE FRECUENCIAS ───────────────────────────────────────────────
    def _tabla_frecuencias(self, indice, serie, hi_umbral) -> TablaFrecuencia:
        """
        Construye la tabla de distribución de frecuencias:
        Rango | Frecuencia | % | Hi acumulada | N empresas acumuladas

        IL, RCI, RP, RA → Hi descendente: Hi[rango] = % empresas con valor ≥ lim_inf[rango]
        IE               → Hi ascendente:  Hi[rango] = % empresas con valor ≤ lim_sup[rango]
        """
        rangos = RANGOS[indice]
        direccion = DIRECCION[indice]
        n = len(serie)

        if n == 0:
            return TablaFrecuencia(indice=indice, filas=[], n_total=0,
                rango_concentrado="Sin datos", indice_recomendado=0.0,
                hi_concentrado=0.0, pct_cumplen=0.0, narrativa="Sin datos.")

        # Contar frecuencias
        filas = []
        for etiq, li, ls in rangos:
            if li is None and ls is not None:
                mask = serie < ls
            elif ls is None and li is not None:
                mask = serie >= li
            elif li is not None and ls is not None:
                mask = (serie >= li) & (serie < ls)
            else:
                mask = pd.Series([True] * n, index=serie.index)
            freq = int(mask.sum())
            filas.append({"etiq": etiq, "li": li, "ls": ls, "freq": freq})

        # Hi acumulada
        if direccion == "mayor":
            # Descendente: empezamos con el 100% y restamos conforme bajamos
            acum = n
            for f in filas:
                f["hi"] = acum / n
                f["n_acum"] = acum
                acum -= f["freq"]
        else:
            # Ascendente (IE): acumulamos de menor a mayor
            acum = 0
            for f in filas:
                acum += f["freq"]
                f["hi"] = acum / n
                f["n_acum"] = acum

        # ── Identificar rango concentrado (método original: umbral fijo) ────
        rango_conc = None
        idx_rec = 0.0
        hi_conc = 0.0

        if direccion == "mayor":
            # Hi baja conforme sube el rango (0-0.99 tiene Hi=100%, 4.50+ tiene Hi=10%).
            # El rango concentrado es el ÚLTIMO rango donde Hi >= umbral antes de caer.
            # Iteramos y guardamos mientras se cumple — cuando deja de cumplirse, paramos.
            for f in filas:
                if f["hi"] >= hi_umbral - 0.005:
                    rango_conc = f["etiq"]
                    hi_conc    = f["hi"]
                    idx_rec    = f["ls"] if f["ls"] is not None else (f["li"] or 0.0)
                else:
                    # Una vez que Hi cae por debajo del umbral, no puede volver a subir
                    # (la Hi descendente es monótona). Paramos aquí.
                    break
        else:
            # IE: Hi sube conforme sube el rango. Tomamos el PRIMER rango que alcanza el umbral.
            for f in filas:
                if f["hi"] >= hi_umbral - 0.005 and rango_conc is None:
                    rango_conc = f["etiq"]
                    hi_conc    = f["hi"]
                    idx_rec    = f["ls"] if f["ls"] is not None else 1.0

        if rango_conc is None:
            ultimo = filas[-1]
            rango_conc = ultimo["etiq"]
            hi_conc = ultimo["hi"]
            idx_rec = ultimo["ls"] if ultimo["ls"] is not None else (ultimo["li"] or 0.0)

        # ── Umbral sugerido por segunda derivada (punto de inflexión real) ──
        # Δ¹[i] = Hi[i+1] - Hi[i]  →  cambio de pendiente entre rangos
        # Δ²[i] = Δ¹[i+1] - Δ¹[i] →  aceleración del cambio
        # El punto de inflexión real es donde |Δ²| es máximo.
        umbral_sug = None
        idx_rec_sug = None
        rango_conc_sug = None

        his = [f["hi"] for f in filas]
        if len(his) >= 3:
            d1 = [his[i+1] - his[i] for i in range(len(his)-1)]
            d2 = [abs(d1[i+1] - d1[i]) for i in range(len(d1)-1)]
            # El índice del máximo |Δ²| corresponde al rango i+1 en filas
            pos_inflexion = d2.index(max(d2)) + 1  # +1 porque d2 empieza en índice 1
            pos_inflexion = min(pos_inflexion, len(filas) - 1)
            f_inf = filas[pos_inflexion]
            umbral_sug = round(f_inf["hi"], 4)
            idx_rec_sug = f_inf["ls"] if f_inf["ls"] is not None else (f_inf["li"] or 0.0)
            rango_conc_sug = f_inf["etiq"]

        # Narrativa (usa el umbral fijo para mantener compatibilidad)
        narrativa = self._narrativa_conclusion(indice, rango_conc, idx_rec, hi_conc, n, direccion)

        # Formatear filas de salida
        filas_out = [
            {
                "Rango": f["etiq"],
                "Frecuencia": f["freq"],
                "%": f"{f['freq']/n:.0%}",
                "Hi": f"{f['hi']:.0%}",
                "Empresas acumuladas": f["n_acum"],
                "_li": f["li"],
                "_ls": f["ls"],
                "_hi_raw": f["hi"],
            }
            for f in filas
        ]

        return TablaFrecuencia(
            indice=indice,
            filas=filas_out,
            n_total=n,
            rango_concentrado=rango_conc,
            indice_recomendado=round(idx_rec, 4),
            hi_concentrado=round(hi_conc, 4),
            pct_cumplen=round(hi_conc, 4),
            narrativa=narrativa,
            umbral_sugerido=umbral_sug,
            indice_recomendado_sugerido=round(idx_rec_sug, 4) if idx_rec_sug is not None else None,
            rango_concentrado_sugerido=rango_conc_sug,
        )

    def _narrativa_conclusion(self, indice, rango_conc, idx_rec, hi_conc, n_total, direccion):
        pct = f"{hi_conc:.0%}"
        n_emp = round(n_total * hi_conc)
        nombre = NARRATIVA[indice]["titulo"].split("(")[0].strip()
        sym = "≥" if direccion == "mayor" else "≤"
        val_str = f"{idx_rec:.0%}" if indice in ("IE", "RP", "RA") else f"{idx_rec:g}"

        if direccion == "mayor":
            return (
                f"En la muestra el {pct} de las empresas se encuentran en el rango "
                f"{rango_conc} o superior del {nombre.lower()}. "
                f"Para efectos del presente análisis se recomienda solicitar un "
                f"{nombre.lower()} mayor o igual a {val_str}, que permita la participación "
                f"del {pct} de los posibles oferentes que pertenecen al sector "
                f"({n_emp} de {n_total} empresas de la muestra)."
            )
        else:
            return (
                f"En la muestra el {pct} de las empresas se encuentran con "
                f"{nombre.lower()} inferior al {val_str}. "
                f"Para efectos del presente análisis se recomienda solicitar un "
                f"{nombre.lower()} menor o igual a {val_str}, que permite la participación "
                f"del {pct} de los posibles oferentes que pertenecen al sector "
                f"({n_emp} de {n_total} empresas de la muestra)."
            )

    # ── MODO B ─────────────────────────────────────────────────────────────
    def _modo_b(self, tabla, objetivo, indice, serie) -> ResultadoIndice:
        direccion = DIRECCION[indice]
        n = len(serie)
        nombre = NARRATIVA[indice]["titulo"].split("(")[0].strip().lower()
        sym = "≥" if direccion == "mayor" else "≤"

        # Rango donde cae el objetivo
        rango_obj = None
        for f in tabla.filas:
            li, ls = f["_li"], f["_ls"]
            if li is None and ls is not None and objetivo < ls:
                rango_obj = f["Rango"]; break
            elif ls is None and li is not None and objetivo >= li:
                rango_obj = f["Rango"]; break
            elif li is not None and ls is not None and li <= objetivo < ls:
                rango_obj = f["Rango"]; break

        # % del sector que cumple el objetivo
        if direccion == "mayor":
            pct_cumple = float((serie >= objetivo).sum()) / n if n > 0 else 0.0
        else:
            pct_cumple = float((serie <= objetivo).sum()) / n if n > 0 else 0.0

        inc = tabla.indice_recomendado
        val_obj = f"{objetivo:.0%}" if indice in ("IE", "RP", "RA") else f"{objetivo:g}"
        val_inc = f"{inc:.0%}" if indice in ("IE", "RP", "RA") else f"{inc:g}"

        if pct_cumple < tabla.hi_concentrado:
            comp = (
                f"El objetivo ({objetivo:g}) es menos exigente que el rango concentrado del sector. "
                f"El {pct_cumple:.0%} del sector cumple este objetivo, vs {tabla.hi_concentrado:.0%} "
                f"que cumpliría el índice recomendado objetivo."
            )
        else:
            comp = (
                f"El objetivo ({objetivo:g}) es más exigente que el rango concentrado del sector. "
                f"Solo el {pct_cumple:.0%} del sector puede cumplirlo."
            )

        narrativa_b = (
            f"Para {nombre}, la empresa establece un índice objetivo de {val_obj}. "
            f"Según la distribución del sector, el {pct_cumple:.0%} de las empresas "
            f"{'supera' if direccion == 'mayor' else 'tiene un endeudamiento inferior a'} este valor. "
            f"{comp} "
            f"▸ Opción inclusiva (rango concentrado): {indice} {sym} {val_inc} "
            f"— el {tabla.hi_concentrado:.0%} del sector puede cumplirlo. "
            f"▸ Opción exigente (objetivo empresa): {indice} {sym} {val_obj} "
            f"— el {pct_cumple:.0%} del sector puede cumplirlo."
        )

        return ResultadoIndice(
            tabla=tabla,
            objetivo_empresa=objetivo,
            rango_objetivo=rango_obj,
            pct_sector_cumple=round(pct_cumple, 4),
            indice_inclusivo=inc,
            narrativa_b=narrativa_b,
        )



    # ── MODO B: SELECCIÓN POR OPTIMIZACIÓN ────────────────────────────────
    def _seleccionar_grupo_modo_b(self, df_sector, indices_objetivo, n_obj, seed):
        """
        Modo B — encuentra las N empresas del sector con perfil más similar
        al de la empresa participante (distancia mínima en espacio normalizado).

        El resultado justifica que existe una muestra representativa del sector
        con un perfil financiero parecido, validando los requisitos del pliego.
        """
        df = self._calcular_indices(df_sector.copy())
        cols_idx = list(RANGOS.keys())
        # Índices obligatorios para calcular distancia (IL e IE siempre disponibles)
        # RCI puede ser NaN para empresas sin deuda financiera — no excluirlas
        cols_obligatorios = ["IL", "IE"]
        cols_disp = [c for c in cols_idx if df[c].notna().sum() > 0]
        df_valido = df.dropna(subset=cols_obligatorios).reset_index(drop=True)

        if len(df_valido) == 0:
            return df_sector.head(n_obj), None

        n_sel = min(n_obj, len(df_valido))

        # Distancia euclidiana usando percentil rank (robusto a outliers)
        # Se calcula por empresa usando solo los índices disponibles (ignora NaN)
        dist_mat = pd.DataFrame(index=df_valido.index)
        for idx in cols_disp:
            col = df_valido[idx]
            # rank solo sobre valores válidos de esa columna
            col_norm = col.rank(pct=True)          # NaN → NaN (no participa)
            n_menores = (col < indices_objetivo[idx]).sum()
            obj_norm  = n_menores / col.notna().sum() if col.notna().sum() > 0 else 0.5
            dist_mat[idx] = (col_norm - obj_norm) ** 2

        # Promedio de dimensiones disponibles por empresa (mean ignora NaN)
        dist = dist_mat.mean(axis=1)
        # Empresas sin ningún índice calculable obtienen distancia máxima (van al final)
        dist = dist.fillna(dist.max() + 1 if dist.notna().any() else 999)
        df_valido["_dist"] = dist
        df_valido = df_valido.sort_values("_dist").reset_index(drop=True)
        grupo = df_valido.head(n_sel).drop(columns=["_dist"], errors="ignore")
        logger.info(f"[ModoB] sector={len(df_sector)} valido={len(df_valido)} n_obj={n_obj} n_sel={n_sel} grupo={len(grupo)}")

        # Métricas por índice
        resumen = {}
        for idx in cols_idx:
            objetivo  = indices_objetivo[idx]
            direccion = DIRECCION[idx]
            serie_g   = grupo[idx].dropna()
            serie_p   = df_valido[idx].dropna()
            promedio  = float(serie_g.mean()) if len(serie_g) > 0 else 0.0
            diff_abs  = abs(promedio - objetivo)
            diff_pct  = (diff_abs / abs(objetivo) * 100) if objetivo != 0 else 0.0

            # % del sector que cumple el índice (para la narrativa del pliego)
            if direccion == "mayor":
                n_pob = int((serie_p >= objetivo).sum())
            else:
                n_pob = int((serie_p <= objetivo).sum())
            pct_pob = n_pob / len(serie_p) if len(serie_p) > 0 else 0.0

            # Empresas del grupo con índice similar (±30%)
            margen = abs(objetivo) * 0.30
            n_sim  = int(((serie_g >= objetivo - margen) & (serie_g <= objetivo + margen)).sum())                      if objetivo != 0 else len(serie_g)

            resumen[idx] = {
                "objetivo":          round(objetivo, 4),
                "promedio_grupo":    round(promedio, 4),
                "diferencia_abs":    round(diff_abs, 4),
                "diferencia_pct":    round(diff_pct, 2),
                "similar":           diff_pct <= 25,
                "n_similares_grupo": n_sim,
                "n_pob_cumple":      n_pob,
                "pct_pob_cumple":    round(pct_pob, 4),
                "n_empresas":        len(grupo),
                "evaluacion":        self._evaluar_similitud(diff_pct, pct_pob),
            }

        return grupo, resumen

    def _evaluar_similitud(self, diff_pct, pct_pob_cumple=0):
        """Evalúa qué tan representativo es el grupo respecto al perfil de la empresa."""
        pct_str = f"{pct_pob_cumple:.0%}"
        if diff_pct <= 10:
            return f"Muy similar — el promedio del grupo es casi igual al índice de la empresa. {pct_str} del sector cumple este valor"
        elif diff_pct <= 25:
            return f"Similar — el grupo tiene un perfil financiero parecido. {pct_str} del sector cumple este valor"
        elif diff_pct <= 50:
            return f"Moderadamente similar — hay diferencia pero el sector tiene empresas cercanas. {pct_str} del sector cumple"
        else:
            if pct_pob_cumple < 0.05:
                return f"Muy diferente — solo el {pct_str} del sector tiene este índice. El requisito puede ser muy exigente para el sector"
            return f"Diferente — el sector tiene un perfil distinto en este índice. {pct_str} del sector cumple"


    def _modo_b_metricas(self, tabla, objetivo, indice, serie, resumen_global) -> ResultadoIndice:
        """Construye ResultadoIndice para Modo B con métricas de cercanía."""
        direccion = DIRECCION[indice]
        n = len(serie)
        nombre = NARRATIVA[indice]["titulo"].split("(")[0].strip().lower()
        sym = "≥" if direccion == "mayor" else "≤"

        # Rango donde cae el objetivo
        rango_obj = None
        for f in tabla.filas:
            li, ls = f["_li"], f["_ls"]
            if li is None and ls is not None and objetivo < ls:
                rango_obj = f["Rango"]; break
            elif ls is None and li is not None and objetivo >= li:
                rango_obj = f["Rango"]; break
            elif li is not None and ls is not None and li <= objetivo < ls:
                rango_obj = f["Rango"]; break

        # % del sector (población) que cumple el objetivo
        pct_cumple = float((serie >= objetivo).sum()) / n if direccion == "mayor" and n > 0 else                      float((serie <= objetivo).sum()) / n if n > 0 else 0.0

        metricas = resumen_global.get(indice, {}) if resumen_global else {}
        promedio_grupo = metricas.get("promedio_grupo")
        diff_abs = metricas.get("diferencia_abs")
        diff_pct = metricas.get("diferencia_pct")
        cumple = metricas.get("cumple_objetivo")
        evaluacion = metricas.get("evaluacion", "")

        val_obj = f"{objetivo:.0%}" if indice in ("IE", "RP", "RA") else f"{objetivo:g}"
        val_prom = f"{promedio_grupo:.0%}" if indice in ("IE", "RP", "RA") else f"{promedio_grupo:g}" if promedio_grupo is not None else "N/D"

        narrativa_b = (
            f"Objetivo para {nombre}: {indice} {sym} {val_obj}. "
            f"El promedio del grupo seleccionado es {val_prom} "
            f"({'cumple' if cumple else 'no cumple'} el objetivo). "
            f"Diferencia: {diff_pct:.1f}%. "
            f"Evaluación: {evaluacion}."
        )

        return ResultadoIndice(
            tabla=tabla,
            objetivo_empresa=objetivo,
            rango_objetivo=rango_obj,
            pct_sector_cumple=round(pct_cumple, 4),
            indice_inclusivo=tabla.indice_recomendado,
            narrativa_b=narrativa_b,
            promedio_grupo=promedio_grupo,
            diferencia_abs=diff_abs,
            diferencia_pct=diff_pct,
            cumple_objetivo=cumple,
        )

# ─────────────────────────────────────────────────────────────────────────────
#  GENERADOR DE GRÁFICAS
# ─────────────────────────────────────────────────────────────────────────────
# Paletas predefinidas — bar_activa, bar_normal, linea, referencia, titulo
PALETAS = {
    "corporativo": {"bar_activa": "#1F4E8C", "bar_normal": "#AED6F1", "linea": "#F39C12", "ref": "#E74C3C", "titulo": "#1F4E8C"},
    "ocean":       {"bar_activa": "#0077B6", "bar_normal": "#90E0EF", "linea": "#F77F00", "ref": "#D62828", "titulo": "#0077B6"},
    "forest":      {"bar_activa": "#2D6A4F", "bar_normal": "#95D5B2", "linea": "#E9C46A", "ref": "#E76F51", "titulo": "#2D6A4F"},
    "slate":       {"bar_activa": "#2C3E50", "bar_normal": "#95A5A6", "linea": "#E67E22", "ref": "#E74C3C", "titulo": "#2C3E50"},
    "wine":        {"bar_activa": "#6D2B6D", "bar_normal": "#C9A0C9", "linea": "#F4A261", "ref": "#E63946", "titulo": "#6D2B6D"},
    "carbon":      {"bar_activa": "#1A1A2E", "bar_normal": "#8B8FA8", "linea": "#FFB347", "ref": "#FF6B6B", "titulo": "#1A1A2E"},
}

# Colores por defecto para compatibilidad
C = {"azul": "#1F4E8C", "azul2": "#2E86C1", "celeste": "#AED6F1",
     "ambar": "#F39C12", "verde": "#27AE60", "rojo": "#E74C3C",
     "conc": "#2E86C1", "normal": "#AED6F1"}


class GeneradorGraficas:
    def __init__(self, resultado: ResultadoAnalisis, paleta: str = "corporativo",
                 incluir_conclusion: bool = True, indices: list | None = None,
                 color_personalizado: dict | None = None):
        self.r = resultado
        # Resolver paleta de colores
        if color_personalizado:
            self.p = {**PALETAS["corporativo"], **color_personalizado}
        else:
            self.p = PALETAS.get(paleta, PALETAS["corporativo"])
        self.incluir_conclusion = incluir_conclusion
        self.indices_incluir = indices or list(RANGOS.keys())

    def _buf(self, fig) -> bytes:
        b = io.BytesIO()
        fig.savefig(b, format="png", bbox_inches="tight", dpi=150, facecolor="white")
        plt.close(fig)
        return b.getvalue()

    def _vacia(self, msg) -> bytes:
        fig, ax = plt.subplots(figsize=(5, 3))
        ax.text(0.5, 0.5, msg, ha="center", va="center", transform=ax.transAxes)
        ax.axis("off")
        return self._buf(fig)

    def histograma_indice(self, indice: str) -> bytes:
        res = self.r.resultados.get(indice)
        if not res or not res.tabla.filas:
            return self._vacia(f"Sin datos: {indice}")

        tabla = res.tabla
        nav = NARRATIVA[indice]
        etq = [f["Rango"] for f in tabla.filas]
        freq = [f["Frecuencia"] for f in tabla.filas]
        hi_vals = [f["_hi_raw"] for f in tabla.filas]
        colores = [self.p["bar_activa"] if f["Rango"] == tabla.rango_concentrado else self.p["bar_normal"]
                   for f in tabla.filas]

        if self.incluir_conclusion:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5),
                                            gridspec_kw={"width_ratios": [2, 1]})
        else:
            fig, ax1 = plt.subplots(1, 1, figsize=(10, 5))
            ax2 = None
        fig.suptitle(nav["titulo"], fontsize=12, fontweight="bold", color=self.p["titulo"], y=1.01)

        # Histograma
        bars = ax1.bar(range(len(etq)), freq, color=colores, edgecolor="white", linewidth=0.8)
        ax1.set_xticks(range(len(etq)))
        ax1.set_xticklabels(etq, rotation=35, ha="right", fontsize=9)
        ax1.set_ylabel("Número de empresas", fontsize=10)
        ax1.set_title("Distribución de frecuencias", fontsize=11)
        ax1.grid(axis="y", alpha=0.3, linestyle="--")
        ax1.set_facecolor("#F8FBFF")
        for sp in ["top", "right"]:
            ax1.spines[sp].set_visible(False)
        for bar, f in zip(bars, freq):
            if f > 0:
                ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.15,
                         str(f), ha="center", fontsize=8, color=self.p["titulo"])

        # Hi en eje secundario
        ax1b = ax1.twinx()
        ax1b.plot(range(len(etq)), [h * 100 for h in hi_vals],
                  color=self.p["linea"], linewidth=2, marker="o", markersize=4, label="Hi (%)")
        ax1b.axhline(self.r.config.hi_umbral * 100, color=self.p["ref"],
                     linewidth=1.5, linestyle="--", alpha=0.7,
                     label=f"Umbral {self.r.config.hi_umbral:.0%}")
        ax1b.set_ylabel("Hi acumulada (%)", fontsize=10)
        ax1b.set_ylim(0, 115)
        ax1b.legend(fontsize=8, loc="upper right")

        # Línea objetivo empresa (modo B)
        if res.objetivo_empresa is not None:
            pos = self._pos_rango(tabla, res.objetivo_empresa)
            ax1.axvline(pos, color=self.p["ref"], linewidth=2, linestyle=":",
                        label=f"Objetivo: {res.objetivo_empresa:g}")
            ax1.legend(fontsize=8)

        # Panel de conclusión (solo si se incluye)
        if ax2 is None:
            plt.tight_layout()
            return self._buf(fig)
        ax2.axis("off")
        ax2.set_facecolor("#F8FBFF")
        y = 0.97

        def txt(texto, y, size=9, bold=False, color="#1A252F"):
            ax2.text(0.05, y, texto, fontsize=size, fontweight="bold" if bold else "normal",
                     color=color, transform=ax2.transAxes, va="top")
            return y - 0.07 * (1 + texto.count("\n") * 0.5)

        y = txt("Conclusión", y, size=11, bold=True, color=self.p["titulo"])
        for linea in self._wrap(tabla.narrativa, 42):
            y = txt(linea, y, size=8.5)
        y -= 0.04
        sym = "≥" if DIRECCION[indice] == "mayor" else "≤"
        y = txt(f"Índice recomendado:", y, bold=True, color=self.p["bar_activa"])
        y = txt(f"  {indice} {sym} {tabla.indice_recomendado:g}", y, size=13, bold=True, color="#27AE60")
        y = txt(f"  Cubre el {tabla.pct_cumplen:.0%} del sector", y, size=9, color=self.p["bar_activa"])

        if res.objetivo_empresa is not None:
            y -= 0.04
            y = txt("Objetivo empresa:", y, bold=True, color=self.p["ref"])
            y = txt(f"  {indice} {sym} {res.objetivo_empresa:g}", y, size=12, bold=True, color=C["rojo"])
            y = txt(f"  Cubre el {res.pct_sector_cumple:.0%} del sector", y, size=9, color=C["rojo"])

        fig.tight_layout()
        return self._buf(fig)

    def distribucion_departamentos(self) -> bytes:
        pob = self.r.tabla_poblacion
        mue = self.r.tabla_muestra
        pob = pob[pob["Departamento"] != "Total general"]
        mue = mue[mue["Departamento"] != "Total general"]
        merged = pob.merge(mue, on="Departamento", how="left", suffixes=(" P", " M")).fillna(0)
        merged = merged.sort_values("Número de Empresas P", ascending=True)

        fig, ax = plt.subplots(figsize=(10, max(4, len(merged) * 0.4)))
        y = range(len(merged))
        ax.barh(y, merged["Número de Empresas P"], color=C["celeste"], label="Población", alpha=0.85)
        ax.barh(y, merged["Número de Empresas M"], color=C["azul2"], label="Muestra", alpha=0.9)
        ax.set_yticks(list(y))
        ax.set_yticklabels(merged["Departamento"].tolist(), fontsize=9)
        ax.set_title(f"Distribución geográfica — Población ({self.r.n_poblacion}) vs Muestra ({self.r.n_muestra})",
                     fontsize=11, fontweight="bold", color=self.p["titulo"])
        ax.legend(fontsize=9)
        ax.grid(axis="x", alpha=0.3, linestyle="--")
        ax.set_facecolor("#F8FBFF")
        for sp in ["top", "right"]:
            ax.spines[sp].set_visible(False)
        fig.tight_layout()
        return self._buf(fig)

    def panel_resumen(self) -> bytes:
        indices = list(RANGOS.keys())
        recom = [self.r.resultados[i].tabla.indice_recomendado for i in indices]
        pcts = [self.r.resultados[i].tabla.pct_cumplen * 100 for i in indices]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle("Resumen de Índices Recomendados",
                     fontsize=13, fontweight="bold", color=self.p["titulo"])

        bars1 = ax1.bar(indices, recom, color=C["conc"], edgecolor="white")
        ax1.set_title("Valor recomendado por índice", fontsize=11)
        ax1.set_ylabel("Valor", fontsize=10)
        ax1.grid(axis="y", alpha=0.3, linestyle="--")
        ax1.set_facecolor("#F8FBFF")
        for sp in ["top", "right"]:
            ax1.spines[sp].set_visible(False)
        for bar, v, idx in zip(bars1, recom, indices):
            sym = "≥" if DIRECCION[idx] == "mayor" else "≤"
            ax1.text(bar.get_x() + bar.get_width() / 2,
                     bar.get_height() + max(recom) * 0.02,
                     f"{sym}{v:g}", ha="center", fontsize=10,
                     fontweight="bold", color=self.p["titulo"])

        bars2 = ax2.bar(indices, pcts, color=C["verde"], edgecolor="white")
        ax2.axhline(self.r.config.hi_umbral * 100, color=C["rojo"],
                    linewidth=1.5, linestyle="--", label=f"Umbral {self.r.config.hi_umbral:.0%}")
        ax2.set_title("% del sector que cumple cada índice", fontsize=11)
        ax2.set_ylabel("%", fontsize=10)
        ax2.set_ylim(0, 115)
        ax2.legend(fontsize=9)
        ax2.grid(axis="y", alpha=0.3, linestyle="--")
        ax2.set_facecolor("#F8FBFF")
        for sp in ["top", "right"]:
            ax2.spines[sp].set_visible(False)
        for bar, p in zip(bars2, pcts):
            ax2.text(bar.get_x() + bar.get_width() / 2,
                     bar.get_height() + 1.5,
                     f"{p:.0f}%", ha="center", fontsize=9, color=self.p["titulo"])

        fig.tight_layout()
        return self._buf(fig)

    def todas_como_zip(self) -> bytes:
        import zipfile
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("00_departamentos.png", self.distribucion_departamentos())
            zf.writestr("00_resumen_indices.png", self.panel_resumen())
            for idx in RANGOS:
                if idx in self.indices_incluir:
                    zf.writestr(f"{idx}_frecuencias.png", self.histograma_indice(idx))
        return buf.getvalue()

    @staticmethod
    def _wrap(texto, ancho):
        palabras = texto.split()
        lineas, linea = [], ""
        for p in palabras:
            if len(linea) + len(p) + 1 > ancho:
                if linea: lineas.append(linea)
                linea = p
            else:
                linea = (linea + " " + p).strip()
        if linea: lineas.append(linea)
        return lineas

    @staticmethod
    def _pos_rango(tabla, valor):
        for i, f in enumerate(tabla.filas):
            li, ls = f["_li"], f["_ls"]
            if li is None and valor < (ls or 0): return i
            elif ls is None and valor >= (li or 0): return i
            elif li is not None and ls is not None and li <= valor < ls: return i
        return len(tabla.filas) - 1


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORTADOR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

class ExportadorExcel:
    AZ = "1F4E8C"; AZ2 = "2E86C1"; AZC = "D6EAF8"
    VD = "D5F5E3"; AM = "FEF9E7"; RJ = "FADBD8"
    BL = "FFFFFF"; GR = "F2F3F4"

    def __init__(self, resultado: ResultadoAnalisis):
        self.r = resultado

    def generar(self) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._resumen(wb)
        for idx in RANGOS:
            self._hoja_indice(wb, idx)
        self._conclusiones(wb)
        self._muestra(wb)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _th(self, c, color=None):
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        c.fill = PatternFill("solid", fgColor=color or self.AZ)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        s = Side(style="thin", color="CCCCCC")
        c.border = Border(top=s, bottom=s, left=s, right=s)

    def _td(self, c, shade=False, bold=False, bg=None, align="left"):
        c.fill = PatternFill("solid", fgColor=bg or (self.GR if shade else self.BL))
        c.font = Font(name="Calibri", size=10, bold=bold)
        s = Side(style="thin", color="DDDDDD")
        c.border = Border(top=s, bottom=s, left=s, right=s)
        c.alignment = Alignment(horizontal=align, vertical="center")

    def _aw(self, ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 55)

    def _resumen(self, wb):
        ws = wb.create_sheet("Resumen")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:D1")
        ws["A1"].value = "FINANALYTICS — Análisis de Sector Financiero"
        ws["A1"].font = Font(bold=True, size=15, color=self.AZ, name="Calibri")
        ws.row_dimensions[1].height = 30

        r = self.r
        params = [
            ("Modo", "A — Objetivo del sector" if r.config.modo == "A" else "B — Por objetivo de empresa"),
            ("CIIUs", ", ".join(r.config.ciius)),
            ("% muestra", f"{r.config.porcentaje_muestra:.0%}"),
            ("Umbral Hi", f"{r.config.hi_umbral:.0%}"),
            ("Año datos", str(r.config.año_datos or "No especificado")),
            ("Población", str(r.n_poblacion)),
            ("Muestra", str(r.n_muestra)),
        ]
        for i, (k, v) in enumerate(params, 3):
            self._td(ws.cell(i, 1, k), bold=True, shade=(i%2==0))
            self._td(ws.cell(i, 2, v), shade=(i%2==0))

        fila = len(params) + 5
        ws.cell(fila, 1).value = "DISTRIBUCIÓN GEOGRÁFICA"
        ws.cell(fila, 1).font = Font(bold=True, size=12, color=self.AZ, name="Calibri")
        fila += 1
        for j, h in enumerate(["Departamento", "Población", "Muestra", "% muestra"], 1):
            self._th(ws.cell(fila, j, h))

        pob = r.tabla_poblacion[r.tabla_poblacion["Departamento"] != "Total general"]
        mue = r.tabla_muestra[r.tabla_muestra["Departamento"] != "Total general"]
        mg = pob.merge(mue, on="Departamento", how="left", suffixes=(" P"," M")).fillna(0)
        for i, row in enumerate(mg.itertuples(index=False), fila+1):
            shade = (i%2==0)
            ws.cell(i,1,row[0]); ws.cell(i,2,int(row[1])); ws.cell(i,3,int(row[2]))
            ws.cell(i,4, f"{row[2]/row[1]:.1%}" if row[1]>0 else "0%")
            for j in range(1,5): self._td(ws.cell(i,j), shade=shade, align="right" if j>1 else "left")
        self._aw(ws)

    def _hoja_indice(self, wb, indice):
        nav = NARRATIVA[indice]
        res = self.r.resultados[indice]
        tabla = res.tabla
        ws = wb.create_sheet(indice)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        ws["A1"].value = nav["titulo"]
        ws["A1"].font = Font(bold=True, size=13, color=self.AZ, name="Calibri")
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:F3")
        ws["A2"].value = nav["intro"]
        ws["A2"].font = Font(size=10, name="Calibri")
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 55

        ws["A4"].value = f"Fórmula: {nav['formula']}"
        ws["A4"].font = Font(bold=True, size=10, color=self.AZ2, name="Calibri")

        fila = 6
        for j, h in enumerate(["Rango","Frecuencia","%","Hi","Empresas acumuladas"], 1):
            self._th(ws.cell(fila, j, h))
        ws.row_dimensions[fila].height = 26

        for i, f in enumerate(tabla.filas, fila+1):
            conc = f["Rango"] == tabla.rango_concentrado
            bg = self.AZC if conc else None
            ws.cell(i,1,f["Rango"]); ws.cell(i,2,f["Frecuencia"])
            ws.cell(i,3,f["%"]); ws.cell(i,4,f["Hi"]); ws.cell(i,5,f["Empresas acumuladas"])
            for j in range(1,6):
                self._td(ws.cell(i,j), shade=(i%2==0) and not conc,
                         bg=bg, align="right" if j>1 else "left", bold=conc)

        ft = fila+len(tabla.filas)+1
        ws.cell(ft,1,"TOTAL"); ws.cell(ft,2,tabla.n_total); ws.cell(ft,3,"100%")
        for j in range(1,4): self._th(ws.cell(ft,j), color=self.AZ2)

        fc = ft+2
        ws.cell(fc,1).value = "CONCLUSIÓN"
        ws.cell(fc,1).font = Font(bold=True, size=11, color=self.AZ, name="Calibri")
        fc += 1
        ws.merge_cells(f"A{fc}:F{fc+2}")
        ws[f"A{fc}"].value = tabla.narrativa
        ws[f"A{fc}"].alignment = Alignment(wrap_text=True)
        ws[f"A{fc}"].font = Font(size=10, name="Calibri")
        ws.row_dimensions[fc].height = 65

        fi = fc+4
        sym = "≥" if DIRECCION[indice]=="mayor" else "≤"
        ws.cell(fi,1,"Índice recomendado:").font = Font(bold=True, size=11, name="Calibri")
        ws.cell(fi,2, f"{indice} {sym} {tabla.indice_recomendado:g}").font = Font(
            bold=True, size=14, color=self.AZ2, name="Calibri")
        ws.cell(fi+1,1,"Empresas que lo cumplen:").font = Font(size=10, name="Calibri")
        ws.cell(fi+1,2, f"{tabla.pct_cumplen:.0%} de la muestra").font = Font(
            size=10, color=self.AZ2, name="Calibri")

        if res.objetivo_empresa is not None:
            fb = fi+3
            ws.cell(fb,1,"ANÁLISIS POR OBJETIVO DE EMPRESA").font = Font(bold=True,size=11,color="993333",name="Calibri")
            ws.cell(fb+1,1,"Objetivo empresa:").font = Font(size=10,bold=True,name="Calibri")
            ws.cell(fb+1,2, f"{indice} {sym} {res.objetivo_empresa:g}").font = Font(bold=True,size=12,color="993333",name="Calibri")
            ws.cell(fb+2,1,"% del sector que cumple:").font = Font(size=10,name="Calibri")
            ws.cell(fb+2,2, f"{res.pct_sector_cumple:.0%}").font = Font(size=10,color="993333",name="Calibri")
            ws.cell(fb+3,1,"Opción inclusiva (rango concentrado):").font = Font(size=10,bold=True,name="Calibri")
            ws.cell(fb+3,2, f"{indice} {sym} {res.indice_inclusivo:g} — {tabla.hi_concentrado:.0%} del sector").font = Font(size=10,color=self.AZ2,name="Calibri")
            if res.narrativa_b:
                fn = fb+5
                ws.merge_cells(f"A{fn}:F{fn+2}")
                ws[f"A{fn}"].value = res.narrativa_b
                ws[f"A{fn}"].alignment = Alignment(wrap_text=True)
                ws[f"A{fn}"].font = Font(size=10,name="Calibri")
                ws.row_dimensions[fn].height = 75
        self._aw(ws)

    def _conclusiones(self, wb):
        ws = wb.create_sheet("Conclusiones")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:F1")
        ws["A1"].value = "RESUMEN DE ÍNDICES RECOMENDADOS"
        ws["A1"].font = Font(bold=True, size=14, color=self.AZ, name="Calibri")
        ws.row_dimensions[1].height = 26
        for j, h in enumerate(["Índice","Descripción","Valor recomendado","Dirección","% cumple","Rango concentrado"],1):
            self._th(ws.cell(3,j,h))
        for i, (idx, res) in enumerate(self.r.resultados.items(), 4):
            sym = "≥" if DIRECCION[idx]=="mayor" else "≤"
            ws.cell(i,1,idx); ws.cell(i,2,NARRATIVA[idx]["titulo"])
            ws.cell(i,3, f"{idx} {sym} {res.tabla.indice_recomendado:g}")
            ws.cell(i,4,"Mayor es mejor" if DIRECCION[idx]=="mayor" else "Menor es mejor")
            ws.cell(i,5, f"{res.tabla.pct_cumplen:.0%}")
            ws.cell(i,6, res.tabla.rango_concentrado)
            for j in range(1,7): self._td(ws.cell(i,j), shade=(i%2==0), align="center" if j in(1,3,5) else "left")
            ws.cell(i,3).font = Font(bold=True,color=self.AZ2,name="Calibri",size=11)
        self._aw(ws)

    def _muestra(self, wb):
        ws = wb.create_sheet("Empresas de la Muestra")
        ws.sheet_view.showGridLines = False
        df = self.r.df_muestra
        cols = [COL["nit"],COL["nombre"],COL["ciiu"],COL["dpto"],COL["ciudad"],"IL","IE","RCI","RP","RA","CT"]
        cols = [c for c in cols if c in df.columns]
        df_out = df[cols].copy()
        for j,h in enumerate(cols,1): self._th(ws.cell(1,j,h))
        ws.freeze_panes = "A2"
        for i, row in enumerate(df_out.itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                v = "" if (val is None or (isinstance(val, float) and np.isnan(val))) else val
                c = ws.cell(i,j,v)
                self._td(c, shade=(i%2==0), align="right" if isinstance(v,(int,float)) else "left")
                if isinstance(v, float): c.number_format = "0.0000"
        self._aw(ws)